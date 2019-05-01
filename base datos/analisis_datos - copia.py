from openpyxl import load_workbook, utils

class Products:

    def __init__(self):
        self.sku = ""
        self.nombre = ""
        self.grupos_productores= ""
        self.stock_minimo = ""
        self.stock_minimo_lotes = ""
        self.duracion_esperada_lotes = ""


lista = []
productos = []
def cargar_base(archivo, inicio, termino):
    wb = load_workbook(archivo)
    ws3 = wb["Análisis"]
    for x in range(inicio, termino+1):
        producto = Products()
        producto.sku = ws3["{}{}".format("A", x)].value
        producto.nombre = ws3["{}{}".format("B", x)].value
        producto.grupos_productores = ws3["{}{}".format("J", x)].value
        producto.stock_minimo = ws3["{}{}".format("M", x)].value
        producto.stock_minimo_lotes = ws3["{}{}".format("N", x)].value
        producto.duracion_produccion = ws3["{}{}".format("I", x)].value
        producto.duracion_esperada_lotes = ws3["{}{}".format("Q", x)].value
        productos.append(producto)
    return productos
diccionario = {}
def cargar_diccionario(archivo, inicio, termino):
    wb = load_workbook(archivo)
    ws3 = wb["Análisis"]
    for x in range(inicio, termino+1):
        diccionario[ws3["{}{}".format("A", x)].value] = {}
        diccionario[ws3["{}{}".format("A", x)].value]["nombre"] = ws3["{}{}".format("B", x)].value
        diccionario[ws3["{}{}".format("A", x)].value]["productores"]  = ws3["{}{}".format("J", x)].value
        diccionario[ws3["{}{}".format("A", x)].value]["stock_minimo"]  = ws3["{}{}".format("M", x)].value
        diccionario[ws3["{}{}".format("A", x)].value]["stock_minimo_lotes"]  = ws3["{}{}".format("N", x)].value
        diccionario[ws3["{}{}".format("A", x)].value]["duracion_produccion"] = ws3["{}{}".format("I", x)].value
        diccionario[ws3["{}{}".format("A", x)].value]["duracion_esperada_lotes"]  = ws3["{}{}".format("Q", x)].value
    return productos


cargar_base("basedatos.xlsx",2,41)
cargar_diccionario("basedatos.xlsx",2,41)
for i in productos:
    print(i.nombre)
print(diccionario)
